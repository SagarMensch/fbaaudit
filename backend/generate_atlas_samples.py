"""
Generate Sample Atlas Invoice and Annexure
==========================================
Creates realistic Indian freight documents for testing the Atlas Freight Audit Platform.

This generates:
1. A consolidated invoice PDF (1 page)
2. An Excel annexure with 25 LR line items that SUM to the invoice total

The totals will MATCH so reconciliation passes.
"""

import sys
sys.path.insert(0, '.')

from datetime import datetime, timedelta
import random
import os

# Import ReportLab for PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# Import openpyxl for Excel
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ============================================================================
# CONFIGURATION - Realistic Indian Freight Data
# ============================================================================

VENDOR = {
    "name": "SHARMA TRANSPORT CORPORATION",
    "address": "Plot No. 45, Transport Nagar, Bhiwandi, Thane - 421302",
    "gstin": "27AABCS1234E1ZV",
    "pan": "AABCS1234E",
    "contact": "+91 98765 43210",
    "email": "accounts@sharmatransport.in"
}

BUYER = {
    "name": "INDOCO REMEDIES LIMITED",
    "address": "B-20, MIDC Industrial Area, Waluj, Aurangabad - 431136",
    "gstin": "27AAACI0380C1Z3",
    "state": "MAHARASHTRA",
    "state_code": "27"
}

# Routes with realistic pricing (per kg or per trip)
ROUTES = [
    {"origin": "Mumbai", "destination": "Delhi", "rate_per_kg": 2.8, "distance_km": 1420},
    {"origin": "Mumbai", "destination": "Bangalore", "rate_per_kg": 2.2, "distance_km": 980},
    {"origin": "Mumbai", "destination": "Chennai", "rate_per_kg": 2.5, "distance_km": 1340},
    {"origin": "Mumbai", "destination": "Ahmedabad", "rate_per_kg": 1.8, "distance_km": 520},
    {"origin": "Mumbai", "destination": "Hyderabad", "rate_per_kg": 2.1, "distance_km": 715},
    {"origin": "Mumbai", "destination": "Kolkata", "rate_per_kg": 3.2, "distance_km": 1990},
    {"origin": "Mumbai", "destination": "Pune", "rate_per_kg": 0.9, "distance_km": 150},
    {"origin": "Mumbai", "destination": "Jaipur", "rate_per_kg": 2.6, "distance_km": 1150},
]

VEHICLE_TYPES = ["20FT Container", "32FT MXL", "22FT SXL", "Tata 407", "14FT Closed"]

# ============================================================================
# GENERATE LINE ITEMS
# ============================================================================

def generate_line_items(num_items=25):
    """Generate realistic LR line items"""
    items = []
    base_date = datetime.now() - timedelta(days=15)
    
    for i in range(num_items):
        route = random.choice(ROUTES)
        weight = random.randint(800, 5000)  # kg
        
        # Calculate freight
        base_freight = round(weight * route["rate_per_kg"], 2)
        
        # Add realistic charges
        fuel_surcharge = round(base_freight * random.uniform(0.08, 0.12), 2)  # 8-12% fuel
        handling = random.choice([0, 500, 750, 1000, 1500])  # Hamali charges
        detention = random.choice([0, 0, 0, 250, 500, 750])  # Mostly 0, sometimes charged
        
        total = round(base_freight + fuel_surcharge + handling + detention, 2)
        
        # Generate LR number (realistic Indian format)
        lr_number = f"STC/{base_date.strftime('%m%y')}/{random.randint(10000, 99999)}"
        lr_date = base_date + timedelta(days=random.randint(-10, 0))
        
        items.append({
            "sr_no": i + 1,
            "lr_number": lr_number,
            "lr_date": lr_date.strftime("%d-%m-%Y"),
            "origin": route["origin"],
            "destination": route["destination"],
            "vehicle_type": random.choice(VEHICLE_TYPES),
            "vehicle_number": f"MH-{random.randint(1,50):02d}-{random.choice(['AA','AB','AC','AD'])}-{random.randint(1000,9999)}",
            "weight_kg": weight,
            "base_freight": base_freight,
            "fuel_surcharge": fuel_surcharge,
            "handling": handling,
            "detention": detention,
            "total": total
        })
    
    return items


def calculate_totals(items):
    """Calculate invoice totals with GST"""
    subtotal = sum(item["total"] for item in items)
    
    # For intra-state (same state), use CGST + SGST
    cgst = round(subtotal * 0.025, 2)  # 2.5%
    sgst = round(subtotal * 0.025, 2)  # 2.5%
    igst = 0  # 0 for intra-state
    
    grand_total = round(subtotal + cgst + sgst, 2)
    
    return {
        "subtotal": subtotal,
        "cgst": cgst,
        "sgst": sgst,
        "igst": igst,
        "grand_total": grand_total
    }


# ============================================================================
# CREATE PDF INVOICE
# ============================================================================

def create_invoice_pdf(items, totals, output_path):
    """Create a professional Indian freight invoice PDF"""
    
    doc = SimpleDocTemplate(output_path, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []
    
    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.gray)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold')
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9)
    
    # Header
    elements.append(Paragraph(VENDOR["name"], title_style))
    elements.append(Paragraph(f'{VENDOR["address"]}', subtitle_style))
    elements.append(Paragraph(f'GSTIN: {VENDOR["gstin"]} | PAN: {VENDOR["pan"]} | {VENDOR["contact"]}', subtitle_style))
    elements.append(Spacer(1, 10))
    
    # Invoice Title
    elements.append(Paragraph('<b>TAX INVOICE</b>', ParagraphStyle('InvTitle', fontSize=14, alignment=TA_CENTER, textColor=colors.HexColor('#1a365d'))))
    elements.append(Spacer(1, 10))
    
    # Invoice Details Table
    invoice_number = f"STC/INV/{datetime.now().strftime('%Y%m')}/{random.randint(1000, 9999)}"
    invoice_date = datetime.now().strftime("%d-%m-%Y")
    
    invoice_info = [
        ['Invoice No:', invoice_number, 'Invoice Date:', invoice_date],
        ['Bill To:', BUYER["name"], 'Place of Supply:', BUYER["state"]],
        ['Address:', BUYER["address"][:50], 'State Code:', BUYER["state_code"]],
        ['GSTIN:', BUYER["gstin"], 'Payment Terms:', 'Net 30 Days'],
    ]
    
    info_table = Table(invoice_info, colWidths=[70, 180, 80, 130])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))
    
    # Annexure Reference
    elements.append(Paragraph(f'<b>Annexure:</b> {len(items)} LR Line Items (See attached Excel)', normal_style))
    elements.append(Spacer(1, 10))
    
    # Summary Table
    summary_data = [
        ['Description', 'HSN/SAC', 'Amount (₹)'],
        ['Freight Charges (As per Annexure)', '996511', f'{totals["subtotal"]:,.2f}'],
        ['', '', ''],
        ['CGST @ 2.5%', '', f'{totals["cgst"]:,.2f}'],
        ['SGST @ 2.5%', '', f'{totals["sgst"]:,.2f}'],
        ['', '', ''],
        ['GRAND TOTAL', '', f'₹ {totals["grand_total"]:,.2f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[280, 80, 100])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f4f8')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Amount in Words
    def num_to_words(num):
        # Simplified - just return a formatted string
        return f"Rupees {int(num):,} and {int((num % 1) * 100)} Paise Only"
    
    elements.append(Paragraph(f'<b>Amount in Words:</b> {num_to_words(totals["grand_total"])}', normal_style))
    elements.append(Spacer(1, 20))
    
    # Bank Details
    bank_info = [
        ['BANK DETAILS'],
        ['Bank Name: HDFC Bank Ltd'],
        ['Account No: 50100123456789'],
        ['IFSC Code: HDFC0001234'],
        ['Branch: Bhiwandi, Thane'],
    ]
    bank_table = Table(bank_info, colWidths=[250])
    bank_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.gray),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(bank_table)
    elements.append(Spacer(1, 30))
    
    # Signature
    sig_data = [
        ['For ' + VENDOR["name"]],
        [''],
        [''],
        ['Authorised Signatory'],
    ]
    sig_table = Table(sig_data, colWidths=[200])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    
    # Right align signature
    elements.append(Table([[sig_table]], colWidths=[460]))
    
    # Build PDF
    doc.build(elements)
    print(f"✅ Invoice PDF created: {output_path}")
    return invoice_number


# ============================================================================
# CREATE EXCEL ANNEXURE
# ============================================================================

def create_annexure_excel(items, totals, output_path, invoice_number):
    """Create Excel annexure with all LR line items"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annexure"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_format = '#,##0.00'
    
    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = f"ANNEXURE TO INVOICE: {invoice_number}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:L2')
    ws['A2'] = f"Vendor: {VENDOR['name']} | Total LRs: {len(items)}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Sr.', 'LR Number', 'LR Date', 'Origin', 'Destination', 
               'Vehicle Type', 'Vehicle No.', 'Weight (Kg)', 
               'Base Freight', 'Fuel Surcharge', 'Handling', 'Detention', 'Total']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows
    for row_idx, item in enumerate(items, 5):
        ws.cell(row=row_idx, column=1, value=item['sr_no']).border = border
        ws.cell(row=row_idx, column=2, value=item['lr_number']).border = border
        ws.cell(row=row_idx, column=3, value=item['lr_date']).border = border
        ws.cell(row=row_idx, column=4, value=item['origin']).border = border
        ws.cell(row=row_idx, column=5, value=item['destination']).border = border
        ws.cell(row=row_idx, column=6, value=item['vehicle_type']).border = border
        ws.cell(row=row_idx, column=7, value=item['vehicle_number']).border = border
        
        weight_cell = ws.cell(row=row_idx, column=8, value=item['weight_kg'])
        weight_cell.border = border
        weight_cell.number_format = '#,##0'
        
        for col, key in enumerate(['base_freight', 'fuel_surcharge', 'handling', 'detention', 'total'], 9):
            cell = ws.cell(row=row_idx, column=col, value=item[key])
            cell.border = border
            cell.number_format = money_format
    
    # Totals row
    total_row = len(items) + 5
    ws.cell(row=total_row, column=8, value="SUBTOTAL:").font = Font(bold=True)
    
    for col, key in enumerate(['base_freight', 'fuel_surcharge', 'handling', 'detention', 'total'], 9):
        cell = ws.cell(row=total_row, column=col, value=sum(item[key] for item in items))
        cell.font = Font(bold=True)
        cell.number_format = money_format
        cell.border = border
    
    # Tax summary
    ws.cell(row=total_row + 2, column=12, value="CGST @ 2.5%:").font = Font(bold=True)
    ws.cell(row=total_row + 2, column=13, value=totals['cgst']).number_format = money_format
    
    ws.cell(row=total_row + 3, column=12, value="SGST @ 2.5%:").font = Font(bold=True)
    ws.cell(row=total_row + 3, column=13, value=totals['sgst']).number_format = money_format
    
    ws.cell(row=total_row + 4, column=12, value="GRAND TOTAL:").font = Font(bold=True, size=12)
    grand_cell = ws.cell(row=total_row + 4, column=13, value=totals['grand_total'])
    grand_cell.font = Font(bold=True, size=12)
    grand_cell.number_format = money_format
    
    # Column widths
    widths = [5, 18, 12, 12, 12, 14, 14, 10, 12, 12, 10, 10, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    wb.save(output_path)
    print(f"✅ Excel Annexure created: {output_path}")


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    print("="*60)
    print("GENERATING ATLAS SAMPLE INVOICE + ANNEXURE")
    print("="*60)
    
    # Generate data
    items = generate_line_items(25)
    totals = calculate_totals(items)
    
    print(f"\n📊 Generated {len(items)} line items")
    print(f"   Subtotal: ₹{totals['subtotal']:,.2f}")
    print(f"   CGST: ₹{totals['cgst']:,.2f}")
    print(f"   SGST: ₹{totals['sgst']:,.2f}")
    print(f"   GRAND TOTAL: ₹{totals['grand_total']:,.2f}")
    
    # Create files
    pdf_path = "Atlas_Sample_Invoice.pdf"
    excel_path = "Atlas_Sample_Annexure.xlsx"
    
    invoice_number = create_invoice_pdf(items, totals, pdf_path)
    create_annexure_excel(items, totals, excel_path, invoice_number)
    
    print("\n" + "="*60)
    print("FILES CREATED - Ready for testing!")
    print("="*60)
    print(f"\n📄 PDF Invoice: {pdf_path}")
    print(f"📊 Excel Annexure: {excel_path}")
    print(f"\n💡 Upload both files to Bulk Invoice Upload to test reconciliation")
